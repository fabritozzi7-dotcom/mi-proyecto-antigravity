"""
dux_export.py — Generador de Excel ENC/DET para importación en Dux.

Toma rendiciones de RENDICIONES_LOG (Google Sheets) y produce un .xlsx
con el formato de 28 columnas (A-AB) que exige el sistema contable Dux.

Uso standalone:  python dux_export.py
Uso importado:   from dux_export import exportar_dux_desde_sheets
"""

import logging
from datetime import datetime
from collections import OrderedDict

logger = logging.getLogger(__name__)

# ==========================================
# CONSTANTES Y MAPEOS
# ==========================================

# Código AFIP → Tipo FP en Dux
AFIP_A_TIPO_FP = {
    "001": "F", "006": "F", "011": "F", "051": "F",
    "002": "ND", "007": "ND", "012": "ND", "052": "ND",
    "003": "NC", "008": "NC", "013": "NC", "053": "NC",
}

# Jurisdicción interna (nombre completo de la sheet) → string que DUX acepta.
# DUX v4 only officially accepts "CABA" and "BS AS" for iibb provincia columns.
# Other provinces pass through as-is (non-blocking) but trigger a validation warning.
JURISDICCION_A_DUX = {
    "CABA": "CABA",
    "CAPITAL FEDERAL": "CABA",
    "CIUDAD AUTONOMA DE BUENOS AIRES": "CABA",
    "BUENOS AIRES": "BS AS",
    "BS AS": "BS AS",
    "PROVINCIA DE BUENOS AIRES": "BS AS",
    "CORDOBA": "CORDOBA",
    "CBA": "CORDOBA",
    "MENDOZA": "MENDOZA",
    "MZA": "MENDOZA",
    "SAN LUIS": "SAN LUIS",
    "SL": "SAN LUIS",
}

# Set of jurisdiction values that DUX v4 officially accepts.
JURISDICCIONES_DUX_VALIDAS = {"CABA", "BS AS", "CORDOBA", "MENDOZA", "SAN LUIS"}

# Jurisdicciones en las que Expoconsult esta inscripta en IIBB.
# Percepciones IIBB de jurisdicciones NO inscriptas se reclasifican a No Gravado.
JURISDICCIONES_INSCRIPTAS = {"CABA", "CAPITAL FEDERAL", "CIUDAD AUTONOMA DE BUENOS AIRES",
                             "BUENOS AIRES", "BS AS", "PROVINCIA DE BUENOS AIRES",
                             "CORDOBA", "CBA", "MENDOZA", "MZA", "SAN LUIS", "SL"}

# Headers de RENDICIONES_LOG → clave interna (por posición 0-30)
# Maps RENDICIONES_LOG columns by position (0-indexed) to internal keys.
# Must match the REAL production header order exactly.
# Used by fila_a_dict() for positional row parsing.
HEADER_MAP = [
    "id_operacion",        # 0  (A)  ID Operación
    "fecha",               # 1  (B)  Fecha
    "usuario",             # 2  (C)  Usuario
    "oficina",             # 3  (D)  Oficina
    "numero_carpeta",      # 4  (E)  Número de Carpeta (Obligatorio)
    "tipo_operacion",      # 5  (F)  Tipo de Operación
    "cliente",             # 6  (G)  Cliente
    "concepto",            # 7  (H)  Concepto
    "monto_concepto",      # 8  (I)  Monto Concepto
    "factura_tipo",        # 9  (J)  factura_tipo
    "codigo_afip",         # 10 (K)  Código AFIP
    "sucursal",            # 11 (L)  Sucursal
    "numero_factura",      # 12 (M)  Número_de_factura
    "n_comprobante",       # 13 (N)  N°Comprobante
    "proveedor_validado",  # 14 (O)  Proveedor_Validado
    "cuit_proveedor",      # 15 (P)  Cuit_Proveedor_AI
    "neto_gravado",        # 16 (Q)  Gravado
    "no_gravado",          # 17 (R)  No_Gravado
    "iva_21",              # 18 (S)  IVA_21 (VALOR IVA SOBRE VALOR GRAVADO)
    "iva_105",             # 19 (T)  IVA_10_5
    "iva_27",              # 20 (U)  IVA_27
    "perc_iva",            # 21 (V)  Percepción_IVA
    "perc_ganancias",      # 22 (W)  Percepción_Ganancia
    "perc_iibb",           # 23 (X)  Percepción IIBB
    "jurisdiccion",        # 24 (Y)  Provincia/Jurisdicción
    "monto_total_ticket",  # 25 (Z)  Monto Ticket
    "monto_a_imputar",     # 26 (AA) Monto a Imputar
    "ticket_url",          # 27 (AB) Ticket URL
    "estado",              # 28 (AC) Estado Saldos
    "clave_maestra",       # 29 (AD) Clave Maestra
    "observaciones",       # 30 (AE) Observaciones
    "aviso_mail",          # 31 (AF) Aviso_Mail
    "cuit_cliente",        # 32 (AG) Cuit_Cliente
    "motivo_rechazo",      # 33 (AH) Motivo_Rechazo
    "revisado_por",        # 34 (AI) Revisado_Por
    "perc_iibb_2",         # 35 (AJ) Perc_IIBB_2
    "jurisdiccion_iibb_2", # 36 (AK) Jurisdiccion_IIBB_2
    "perc_municipal",      # 37 (AL) Perc_Municipal
    "jurisdiccion_municipal", # 38 (AM) Jurisdiccion_Municipal
    "fecha_revision",      # 39 (AN) Fecha_Revision
    "perc_iibb_3",         # 40 (AO) Perc_IIBB_3
    "jurisdiccion_iibb_3", # 41 (AP) Jurisdiccion_IIBB_3
]

DUX_HEADERS = [
    "Tipo Renglón",    # A
    "Tipo Factura",    # B
    "Fecha",           # C
    "Tipo FP",         # D
    "Letra FP",        # E
    "Suc. FP",         # F
    "Nro. FP",         # G
    "Concepto",        # H
    "CUIT",            # I
    "Op.",             # J
    "Mon.",            # K
    "Importe",         # L
    "Detalle",         # M
    "idEmpleado",      # N
    "tipo cambio",     # O
    "neto 21",         # P
    "neto 10.5",       # Q
    "neto 27",         # R
    "noGravado",       # S
    "exento",          # T
    "iva21",           # U
    "iva10",           # V
    "iva27",           # W
    "iibb importe 1",  # X
    "iibb provincia 1",# Y
    "iibb importe 2",  # Z
    "iibb provincia 2",# AA
    "iibb importe 3",  # AB
    "iibb provincia 3",# AC
    "percepcion iva",  # AD
    "percepcion ganancias",  # AE
]


# ==========================================
# HELPERS
# ==========================================

def safe_float(value, default=0.0):
    """Convierte a float de forma segura, limpiando $ y comas."""
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def safe_int(value, default=0):
    """Convierte a int, removiendo ceros a la izquierda."""
    if value is None or value == "":
        return default
    try:
        return int(str(value).strip().lstrip("0") or "0")
    except (ValueError, TypeError):
        return default


def resolver_tipo_fp(codigo_afip):
    """Código AFIP → Tipo FP (F, ND, NC). Default F."""
    code = str(codigo_afip or "").strip().zfill(3)
    return AFIP_A_TIPO_FP.get(code, "F")


def resolver_id_tesoreria(nombre_usuario, codigo_empleado_fn=None):
    """Nombre de usuario → ID cuenta tesorería Dux.

    Args:
        nombre_usuario: user name from RENDICIONES_LOG.
        codigo_empleado_fn: callable(nombre) -> int|None, dynamic lookup.
            If None, returns "" (no mapping available).
    """
    if codigo_empleado_fn is None:
        return ""
    code = codigo_empleado_fn(nombre_usuario)
    return str(code) if code is not None else ""


def resolver_jurisdiccion_dux(jurisdiccion_raw):
    """Normaliza una jurisdicción del sistema interno al formato DUX.

    DUX v4 only accepts "CABA" and "BS AS" for iibb provincia columns.
    Other values pass through as-is (non-blocking) for manual review.
    Treats "0", "0.0", "None", "nan" as absence (Google Sheets stores
    empty numeric cells as "0").
    """
    if not jurisdiccion_raw:
        return ""
    s = str(jurisdiccion_raw).strip()
    if s in ("", "0", "0.0", "None", "nan"):
        return ""
    norm = s.upper()
    return JURISDICCION_A_DUX.get(norm, norm)


def es_propia(cuit_cliente, cuits_propios=None):
    """Determina si un comprobante es PROPIA.

    PROPIA when the client (receiver) CUIT matches any of Expoconsult's CUITs.
    Does NOT depend on factura_tipo — only on who the invoice is addressed to.

    Args:
        cuit_cliente: CUIT of the client/receiver from the invoice.
        cuits_propios: list of Expoconsult CUITs (from CONFIG_EMPRESA).
    """
    if not cuit_cliente or not cuits_propios:
        return False
    clean = str(cuit_cliente).replace("-", "").replace(" ", "").strip()
    if not clean:
        return False
    return clean in cuits_propios


def fila_a_dict(row):
    """Convierte una fila (lista) de RENDICIONES_LOG a dict con claves internas."""
    d = {}
    for i, key in enumerate(HEADER_MAP):
        d[key] = row[i] if i < len(row) else ""
    return d


# ==========================================
# LECTURA DESDE GOOGLE SHEETS
# ==========================================

def leer_rendiciones_desde_sheet(client, sheet_id, sheet_name):
    """
    Lee RENDICIONES_LOG y devuelve lista de dicts con claves internas.

    Args:
        client: gspread client autorizado.
        sheet_id: ID del spreadsheet (o None para abrir por nombre).
        sheet_name: Nombre del spreadsheet (fallback si no hay ID).

    Returns:
        Lista de dicts, uno por fila (sin header).
    """
    if sheet_id:
        sh = client.open_by_key(sheet_id)
    else:
        sh = client.open(sheet_name)

    ws = sh.worksheet("RENDICIONES_LOG")
    all_rows = ws.get_all_values(value_render_option="UNFORMATTED_VALUE")

    if len(all_rows) < 2:
        return []

    # Saltar header (fila 0)
    return [fila_a_dict(row) for row in all_rows[1:]]


# ==========================================
# AGRUPACIÓN POR COMPROBANTE
# ==========================================

def _clave_comprobante(rend):
    """Genera clave de agrupación: CUIT + tipo + sucursal + número."""
    cuit = str(rend.get("cuit_proveedor", "")).replace("-", "").strip()
    tipo = str(rend.get("factura_tipo", "")).strip().upper()
    suc = str(rend.get("sucursal", "")).strip()
    num = str(rend.get("numero_factura", "")).strip()
    return f"{cuit}|{tipo}|{suc}|{num}"


def agrupar_por_comprobante(rendiciones):
    """
    Agrupa rendiciones por clave de comprobante.

    Args:
        rendiciones: lista de dicts (salida de leer_rendiciones_desde_sheet
                     o de fila_a_dict).

    Returns:
        OrderedDict {clave: [lista de dicts del grupo]}, preservando orden
        de aparición.
    """
    grupos = OrderedDict()
    for rend in rendiciones:
        clave = _clave_comprobante(rend)
        if not clave or clave == "|||":
            continue
        grupos.setdefault(clave, []).append(rend)
    return grupos


# ==========================================
# GENERACIÓN DE FILAS DUX
# ==========================================

def _construir_enc(grupo, tipo_factura_dux, total_importe, desglose_sumado):
    """Construye la fila ENC a partir de un grupo de rendiciones."""
    primer = grupo[0]
    fecha_raw = str(primer.get("fecha", "")).strip()

    # Normalizar fecha a DD/MM/AAAA
    fecha = fecha_raw
    if "-" in fecha_raw:
        try:
            dt = datetime.strptime(fecha_raw[:10], "%Y-%m-%d")
            fecha = dt.strftime("%d/%m/%Y")
        except ValueError:
            pass

    tipo_fp = resolver_tipo_fp(primer.get("codigo_afip"))
    letra = str(primer.get("factura_tipo", "C")).strip().upper()
    if letra not in ("A", "B", "C", "M"):
        letra = "C"

    suc = safe_int(primer.get("sucursal"))
    nro = safe_int(primer.get("numero_factura"))
    cuit_proveedor = str(primer.get("cuit_proveedor", "")).replace("-", "").replace(" ", "").strip()

    is_propia = tipo_factura_dux == "PROPIA"

    fila = [""] * len(DUX_HEADERS)
    fila[0] = "ENC"                              # A: Tipo Renglón
    fila[1] = tipo_factura_dux                    # B: TERCEROS / PROPIA
    fila[2] = fecha                               # C: Fecha
    fila[3] = tipo_fp                             # D: Tipo FP
    fila[4] = letra                               # E: Letra FP
    fila[5] = suc                                 # F: Suc. FP
    fila[6] = nro                                 # G: Nro. FP
    # H: Concepto — vacío en ENC
    fila[8] = cuit_proveedor                      # I: CUIT — ALWAYS (obligatorio en ENC)
    # J: Op. — vacío en ENC
    fila[10] = "PES"                              # K: Mon.
    fila[11] = round(total_importe, 2)            # L: Importe

    # M: Detalle — "CONSUMIDOR FINAL" si TERCEROS sin CUIT
    if not is_propia and not cuit_proveedor:
        fila[12] = "CONSUMIDOR FINAL"

    # O-AE: Desglose fiscal (solo PROPIA)
    if is_propia:
        neto_total = desglose_sumado.get("neto_gravado", 0.0)
        iva_21 = desglose_sumado.get("iva_21", 0.0)
        iva_105 = desglose_sumado.get("iva_105", 0.0)
        iva_27 = desglose_sumado.get("iva_27", 0.0)
        
        n21 = n105 = n27 = 0.0
        
        if iva_105 > 0 and iva_21 == 0 and iva_27 == 0:
            n105 = neto_total
        elif iva_27 > 0 and iva_21 == 0 and iva_105 == 0:
            n27 = neto_total
        elif iva_21 > 0 and iva_105 == 0 and iva_27 == 0:
            n21 = neto_total
        else:
            if iva_21 == 0 and iva_105 == 0 and iva_27 == 0:
                n21 = neto_total
            else:
                n21 = round(iva_21 / 0.21, 2) if iva_21 > 0 else 0.0
                n105 = round(iva_105 / 0.105, 2) if iva_105 > 0 else 0.0
                n27 = round(iva_27 / 0.27, 2) if iva_27 > 0 else 0.0

        fila[14] = ""                                              # O: tipo cambio
        fila[15] = round(n21, 2)                                   # P: neto 21
        fila[16] = round(n105, 2)                                  # Q: neto 10.5
        fila[17] = round(n27, 2)                                   # R: neto 27
        fila[18] = round(desglose_sumado["no_gravado"], 2)         # S
        fila[19] = round(desglose_sumado["exento"], 2)             # T
        fila[20] = round(desglose_sumado["iva_21"], 2)             # U
        fila[21] = round(desglose_sumado["iva_105"], 2)            # V
        fila[22] = round(desglose_sumado["iva_27"], 2)             # W

        # IIBB: hasta 3 jurisdicciones con importe
        iibb_list = desglose_sumado.get("iibb", [])
        for slot_idx, iibb_entry in enumerate(iibb_list[:3]):
            base_col = 23 + slot_idx * 2  # X/Z/AB = 23/25/27
            fila[base_col] = round(iibb_entry["importe"], 2)
            fila[base_col + 1] = iibb_entry["provincia"]

        fila[29] = round(desglose_sumado["perc_iva"], 2)          # AD
        fila[30] = round(desglose_sumado["perc_ganancias"], 2)    # AE

    return fila


def _construir_det(rend, tipo_factura_dux, codigo_concepto_fn=None, codigo_empleado_fn=None):
    """Construye una fila DET a partir de una rendición individual.

    Args:
        rend: dict with internal keys.
        tipo_factura_dux: "PROPIA" or "TERCEROS" (from group-level decision).
        codigo_concepto_fn: callable(concepto) -> int|None.
        codigo_empleado_fn: callable(usuario) -> int|None.
    """
    fecha_raw = str(rend.get("fecha", "")).strip()
    fecha = fecha_raw
    if "-" in fecha_raw:
        try:
            dt = datetime.strptime(fecha_raw[:10], "%Y-%m-%d")
            fecha = dt.strftime("%d/%m/%Y")
        except ValueError:
            pass

    tipo_fp = resolver_tipo_fp(rend.get("codigo_afip"))
    letra = str(rend.get("factura_tipo", "C")).strip().upper()
    if letra not in ("A", "B", "C", "M"):
        letra = "C"

    suc = safe_int(rend.get("sucursal"))
    nro = safe_int(rend.get("numero_factura"))

    importe = safe_float(rend.get("monto_a_imputar"))
    concepto_interno = str(rend.get("concepto", "")).strip()
    carpeta = str(rend.get("numero_carpeta", "")).strip()
    usuario = str(rend.get("usuario", "")).strip()

    # DUX code lookups
    codigo_concepto = ""
    if codigo_concepto_fn:
        code = codigo_concepto_fn(concepto_interno)
        if code is not None:
            codigo_concepto = code  # int

    id_tesoreria = resolver_id_tesoreria(usuario, codigo_empleado_fn)

    fila = [""] * len(DUX_HEADERS)
    fila[0] = "DET"                               # A
    fila[1] = tipo_factura_dux                     # B
    fila[2] = fecha                                # C
    fila[3] = tipo_fp                              # D
    fila[4] = letra                                # E
    fila[5] = suc                                  # F
    fila[6] = nro                                  # G
    fila[7] = codigo_concepto                      # H: DUX concept code (numeric)
    # I: CUIT — vacío en DET
    fila[9] = carpeta                              # J: Op. (número de carpeta)
    # K: Mon. — vacío en DET
    fila[11] = round(importe, 2)                   # L
    # M: Detalle
    fila[12] = f"Rendición {usuario} - {concepto_interno}" if concepto_interno else ""
    fila[13] = id_tesoreria                        # N: idEmpleado (DUX code)
    # O-AB: vacío en DET

    return fila


def _sumar_desglose(grupo):
    """Suma los desgloses fiscales prorrateados de un grupo de rendiciones.

    IMPORTANT: This function expects dicts already mapped to INTERNAL keys
    via SHEET_KEY_MAP (e.g. "neto_gravado", "perc_iibb", "jurisdiccion").
    If it receives dicts with raw sheet headers, it will silently return 0s.
    The mapping chain is: get_all_records() → SHEET_KEY_MAP → this function.
    """
    totales = {
        "neto_gravado": 0.0,
        "no_gravado": 0.0,
        "exento": 0.0,
        "iva_21": 0.0,
        "iva_105": 0.0,
        "iva_27": 0.0,
        "perc_iva": 0.0,
        "perc_ganancias": 0.0,
    }

    # Acumular IIBB por jurisdicción (normalizada a formato DUX)
    iibb_por_juris = {}

    def _acumular_iibb(importe_key, juris_key, rend):
        importe = safe_float(rend.get(importe_key))
        juris_raw = str(rend.get(juris_key, "")).strip()
        if importe > 0 and juris_raw:
            # Si la jurisdiccion no esta inscripta, reclasificar a No Gravado
            juris_norm = juris_raw.upper().strip()
            if juris_norm not in JURISDICCIONES_INSCRIPTAS:
                totales["no_gravado"] += importe
                logger.info(
                    f"IIBB {juris_raw} no inscripta -> reclasificada a No Gravado (${importe:.2f})"
                )
                return
            provincia_dux = resolver_jurisdiccion_dux(juris_raw)
            iibb_por_juris[provincia_dux] = iibb_por_juris.get(provincia_dux, 0.0) + importe

    for rend in grupo:
        totales["neto_gravado"] += safe_float(rend.get("neto_gravado"))
        totales["no_gravado"] += safe_float(rend.get("no_gravado"))
        totales["iva_21"] += safe_float(rend.get("iva_21"))
        totales["iva_105"] += safe_float(rend.get("iva_105"))
        totales["iva_27"] += safe_float(rend.get("iva_27"))
        totales["perc_iva"] += safe_float(rend.get("perc_iva"))
        totales["perc_ganancias"] += safe_float(rend.get("perc_ganancias"))

        # IIBB slot 1 (primary)
        _acumular_iibb("perc_iibb", "jurisdiccion", rend)
        # IIBB slot 2
        _acumular_iibb("perc_iibb_2", "jurisdiccion_iibb_2", rend)
        # IIBB slot 3
        _acumular_iibb("perc_iibb_3", "jurisdiccion_iibb_3", rend)
        # Percepcion Municipal -> No Gravado en DUX (confirmado por Fabian 04/05).
        # La columna Perc_Municipal en RENDICIONES_LOG se mantiene para deteccion
        # visual y carga manual en DUX con cuenta 30 - PERCEPCIONES MUNICIPALES.
        perc_muni = safe_float(rend.get("perc_municipal"))
        if perc_muni > 0:
            totales["no_gravado"] += perc_muni

    if len(iibb_por_juris) > 3:
        excedentes = sorted(iibb_por_juris.keys())[3:]
        logger.warning(
            f"Grupo tiene {len(iibb_por_juris)} jurisdicciones IIBB. "
            f"DUX solo soporta 3. No se exportan: {excedentes}"
        )

    totales["iibb"] = [
        {"provincia": prov, "importe": imp}
        for prov, imp in sorted(iibb_por_juris.items())
    ]

    return totales


def generar_filas_dux(grupos, cuits_propios=None, codigo_concepto_fn=None,
                      codigo_empleado_fn=None):
    """
    Genera la lista completa de filas ENC/DET para el Excel Dux.

    Args:
        grupos: OrderedDict de {clave: [rendiciones]}, salida de
                agrupar_por_comprobante.
        cuits_propios: list of Expoconsult CUITs for PROPIA detection.
        codigo_concepto_fn: callable(concepto) -> int|None.
        codigo_empleado_fn: callable(usuario) -> int|None.

    Returns:
        Lista de listas (cada sublista = 28 celdas, una fila del Excel).

    Raises:
        ValueError: if sum(DET) differs from ENC by more than $1 for any group.
    """
    filas = []

    for clave, grupo in grupos.items():
        first = grupo[0]
        cuit_cliente = str(first.get("cuit_cliente", "")).replace("-", "").replace(" ", "").strip()
        is_propia_flag = es_propia(cuit_cliente, cuits_propios)
        tipo_factura_dux = "PROPIA" if is_propia_flag else "TERCEROS"

        # Total ENC = suma de montos a imputar de todas las filas del grupo
        total_importe = sum(safe_float(r.get("monto_a_imputar")) for r in grupo)

        # Desglose sumado (solo relevante para PROPIA, pero lo calculamos siempre)
        desglose_sumado = _sumar_desglose(grupo)

        # Fila ENC
        enc = _construir_enc(grupo, tipo_factura_dux, total_importe, desglose_sumado)
        filas.append(enc)

        # Filas DET (una por rendición/carpeta)
        det_importes = []
        for rend in grupo:
            det = _construir_det(rend, tipo_factura_dux, codigo_concepto_fn, codigo_empleado_fn)
            filas.append(det)
            det_importes.append(safe_float(det[11]))

        # Validate sum(DET) == ENC
        sum_det = sum(det_importes)
        diff = abs(total_importe - sum_det)
        if diff > 1.0:
            cuit_prov = str(first.get("cuit_proveedor", "")).strip()
            raise ValueError(
                f"Sum(DET)={sum_det:.2f} != ENC={total_importe:.2f} "
                f"(diff=${diff:.2f}) for {cuit_prov} — aborting"
            )
        elif diff > 0.005:
            # Adjust last DET for rounding
            last_det_idx = len(filas) - 1
            filas[last_det_idx][11] = round(filas[last_det_idx][11] + (total_importe - sum_det), 2)

    return filas


# ==========================================
# VALIDACIÓN PRE-EXPORT
# ==========================================


def validar_rendiciones_para_export(rendiciones, codigo_concepto_fn=None,
                                    codigo_empleado_fn=None, cuits_propios=None):
    """Validates renditions before DUX export.

    Returns:
        (list[dict], list[dict]): (errores, warnings).
        errores are blocking — abort export.
        warnings are informational — show but don't block.
        Each dict has keys: tipo, mensaje, filas_afectadas, accion.
    """
    errors = []
    warnings = []

    # Group errors by type
    sin_cuit = []
    sin_concepto_dux = {}  # concepto -> [row indices/ids]
    sin_empleado_dux = {}  # usuario -> [row indices/ids]
    propia_sin_desglose = []
    sin_carpeta = []

    for i, rend in enumerate(rendiciones):
        row_ref = rend.get("id_operacion", f"fila {i+1}")

        # 1. CUIT proveedor (11 digits)
        cuit = str(rend.get("cuit_proveedor", "")).replace("-", "").replace(" ", "").strip()
        if not cuit or len(cuit) != 11 or not cuit.isdigit():
            sin_cuit.append(f"row {row_ref}: CUIT='{cuit}'")

        # 2. Concepto → codigo DUX
        concepto = str(rend.get("concepto", "")).strip()
        if concepto and codigo_concepto_fn:
            code = codigo_concepto_fn(concepto)
            if code is None:
                sin_concepto_dux.setdefault(concepto, []).append(row_ref)

        # 3. Usuario → idEmpleado DUX
        usuario = str(rend.get("usuario", "")).strip()
        if usuario and codigo_empleado_fn:
            code = codigo_empleado_fn(usuario)
            if code is None:
                sin_empleado_dux.setdefault(usuario, []).append(row_ref)

        # 4. PROPIA sin desglose
        cuit_cliente = str(rend.get("cuit_cliente", "")).replace("-", "").replace(" ", "").strip()
        if es_propia(cuit_cliente, cuits_propios):
            neto = safe_float(rend.get("neto_gravado"))
            no_grav = safe_float(rend.get("no_gravado"))
            if neto == 0 and no_grav == 0:
                propia_sin_desglose.append(f"row {row_ref}")

        # 5. Carpeta
        carpeta = str(rend.get("numero_carpeta", "")).strip()
        if not carpeta:
            sin_carpeta.append(f"row {row_ref}")

    # Build error list
    if sin_cuit:
        errors.append({
            "tipo": "CUIT proveedor inválido",
            "mensaje": f"{len(sin_cuit)} rendiciones con CUIT vacío o inválido (debe ser 11 dígitos)",
            "filas_afectadas": sin_cuit,
            "accion": "Editá la rendición y completá el CUIT manualmente",
        })

    if sin_concepto_dux:
        total = sum(len(v) for v in sin_concepto_dux.values())
        detalles = []
        for conc, rows in sin_concepto_dux.items():
            detalles.append(f'  "{conc}" (rows: {", ".join(str(r) for r in rows[:5])})')
        errors.append({
            "tipo": "Concepto sin código DUX",
            "mensaje": f"{total} rendiciones con concepto sin mapear",
            "filas_afectadas": detalles,
            "accion": "Completá la columna 'concepto_interno' en MAESTRO_CONCEPTOS_DUX",
        })

    if sin_empleado_dux:
        total = sum(len(v) for v in sin_empleado_dux.values())
        detalles = []
        for usr, rows in sin_empleado_dux.items():
            detalles.append(f'  "{usr}" (rows: {", ".join(str(r) for r in rows[:5])})')
        errors.append({
            "tipo": "Usuario sin idEmpleado DUX",
            "mensaje": f"{total} rendiciones con usuario sin código de tesorería",
            "filas_afectadas": detalles,
            "accion": "Completá 'codigo_dux' en la hoja USUARIOS",
        })

    if propia_sin_desglose:
        errors.append({
            "tipo": "Factura PROPIA sin desglose impositivo",
            "mensaje": f"{len(propia_sin_desglose)} facturas PROPIA sin netoGravado ni noGravado",
            "filas_afectadas": propia_sin_desglose,
            "accion": "Editá la rendición y completá los importes de desglose",
        })

    if sin_carpeta:
        errors.append({
            "tipo": "Rendición sin número de carpeta",
            "mensaje": f"{len(sin_carpeta)} rendiciones sin carpeta (col J del DET quedará vacía)",
            "filas_afectadas": sin_carpeta,
            "accion": "Editá la rendición y completá el número de carpeta",
        })

    # WARNING (not blocking): jurisdicciones IIBB que DUX puede no aceptar
    juris_no_estandar = []
    for i, rend in enumerate(rendiciones):
        row_ref = rend.get("id_operacion", f"fila {i+1}")
        for juris_key in ("jurisdiccion", "jurisdiccion_iibb_2", "jurisdiccion_municipal"):
            juris_raw = str(rend.get(juris_key, "")).strip()
            if not juris_raw:
                continue
            juris_dux = resolver_jurisdiccion_dux(juris_raw)
            if juris_dux and juris_dux not in JURISDICCIONES_DUX_VALIDAS:
                juris_no_estandar.append(
                    f"row {row_ref}: {juris_key}='{juris_raw}' -> '{juris_dux}'"
                )

    if juris_no_estandar:
        warnings.append({
            "tipo": "Jurisdicción IIBB no estándar para DUX",
            "mensaje": (
                f"{len(juris_no_estandar)} percepciones con jurisdicción que DUX puede no aceptar. "
                f"DUX v4 solo acepta 'CABA' y 'BS AS'."
            ),
            "filas_afectadas": juris_no_estandar,
            "accion": "Verificá con el equipo de DUX si la jurisdicción es válida o ajustá manualmente.",
        })

    return errors, warnings


# ==========================================
# EXPORTACIÓN A EXCEL
# ==========================================

def exportar_excel_dux(filas, output_path):
    """
    Escribe las filas ENC/DET en un Excel formateado para Dux.

    Args:
        filas: lista de listas (salida de generar_filas_dux).
        output_path: ruta del archivo .xlsx de salida.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Dux Import"

    # — Estilos —
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    enc_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    det_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    number_fmt = '#,##0.00'

    # — Header (fila 1) —
    for col_idx, header in enumerate(DUX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # — Datos (fila 2+) —
    for row_idx, fila in enumerate(filas, start=2):
        is_enc = (fila[0] == "ENC") if fila else False
        row_fill = enc_fill if is_enc else det_fill

        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.fill = row_fill

            # Formato numérico para columnas de importes
            if isinstance(valor, (int, float)) and valor != 0:
                cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    # — Anchos de columna —
    col_widths = {
        1: 14,   # Tipo Renglón
        2: 14,   # Tipo Factura
        3: 12,   # Fecha
        4: 8,    # Tipo FP
        5: 8,    # Letra FP
        6: 8,    # Suc.
        7: 12,   # Nro.
        8: 30,   # Concepto
        9: 14,   # CUIT
        10: 18,  # Op. (carpeta)
        11: 6,   # Mon.
        12: 14,  # Importe
        13: 20,  # Detalle
        14: 12,  # idEmpleado
    }
    for col_idx in range(15, 29):
        col_widths[col_idx] = 14

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # — Freeze panes (fijar header) —
    ws.freeze_panes = "A2"

    # — Autofiltro —
    last_col_letter = get_column_letter(len(DUX_HEADERS))
    last_row = len(filas) + 1
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    wb.save(output_path)
    logger.info(f"Excel Dux exportado: {output_path} ({len(filas)} filas)")


# ==========================================
# PIPELINE COMPLETO
# ==========================================

def exportar_dux_desde_sheets(rendiciones_raw, output_path, cuits_propios=None,
                              codigo_concepto_fn=None, codigo_empleado_fn=None):
    """
    Pipeline completo: recibe lista de dicts (rendiciones), agrupa,
    genera filas ENC/DET y escribe el Excel.

    Args:
        rendiciones_raw: lista de dicts con claves internas (HEADER_MAP).
        output_path: ruta del .xlsx de salida.
        cuits_propios: list of Expoconsult CUITs.
        codigo_concepto_fn: callable(concepto) -> int|None.
        codigo_empleado_fn: callable(usuario) -> int|None.

    Returns:
        int — cantidad de filas generadas (sin contar header).
    """
    grupos = agrupar_por_comprobante(rendiciones_raw)
    filas = generar_filas_dux(grupos, cuits_propios, codigo_concepto_fn, codigo_empleado_fn)
    exportar_excel_dux(filas, output_path)
    return len(filas)


# ==========================================
# BLOQUE DE PRUEBA STANDALONE
# ==========================================

if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO)

    CUITS_PROPIOS = ["30570717630"]
    MOCK_CONCEPTOS = {"FLETE TERRESTRE": 911, "HONORARIOS DESPACHANTE": 5102,
                      "GASTOS GENERALES OF. BS.AS.": 5148}
    MOCK_EMPLEADOS = {"DAVID REQUELME": 319, "FABRICIO DAURIA": 361}

    def mock_concepto_fn(c):
        return MOCK_CONCEPTOS.get(str(c).strip().upper())

    def mock_empleado_fn(u):
        return MOCK_EMPLEADOS.get(str(u).strip().upper())

    counts = {"passed": 0, "failed": 0}

    def check(name, condition, detail=""):
        if condition:
            print(f"  [PASS] {name}" + (f" -- {detail}" if detail else ""))
            counts["passed"] += 1
        else:
            print(f"  [FAIL] {name}" + (f" -- {detail}" if detail else ""))
            counts["failed"] += 1

    # ── Test 1: TERCEROS sin IIBB ────────────────────────────────────
    print("\n=== Test 1: TERCEROS sin percepciones IIBB ===")
    t1 = [{
        "id_operacion": "T1", "fecha": "2026-05-01",
        "usuario": "DAVID REQUELME", "oficina": "BUENOS AIRES",
        "numero_carpeta": "IMP-001", "tipo_operacion": "Importacion",
        "cliente": "Cliente X", "concepto": "Flete terrestre",
        "monto_concepto": 50000, "factura_tipo": "B", "codigo_afip": "006",
        "sucursal": "00010", "numero_factura": "00045678",
        "n_comprobante": "0001000045678", "proveedor_validado": "Si",
        "cuit_proveedor": "30712345678", "cuit_cliente": "",
        "neto_gravado": 0, "no_gravado": 50000.0,
        "iva_21": 0, "iva_105": 0, "iva_27": 0,
        "perc_iva": 0, "perc_ganancias": 0,
        "perc_iibb": 0, "jurisdiccion": "",
        "perc_iibb_2": 0, "jurisdiccion_iibb_2": "",
        "perc_municipal": 0, "jurisdiccion_municipal": "",
        "monto_total_ticket": 50000.0, "monto_a_imputar": 50000.0,
        "ticket_url": "", "estado": "CERRADO",
        "clave_maestra": "", "observaciones": "",
    }]
    g1 = agrupar_por_comprobante(t1)
    f1 = generar_filas_dux(g1, CUITS_PROPIOS, mock_concepto_fn, mock_empleado_fn)
    enc1 = [r for r in f1 if r[0] == "ENC"][0]
    check("ENC is TERCEROS", enc1[1] == "TERCEROS", enc1[1])
    check("IIBB slot 1 empty", enc1[23] == "" and enc1[24] == "", f"X={enc1[23]} Y={enc1[24]}")
    check("IIBB slot 2 empty", enc1[25] == "" and enc1[26] == "", f"Z={enc1[25]} AA={enc1[26]}")
    check("IIBB slot 3 empty", enc1[27] == "" and enc1[28] == "", f"AB={enc1[27]} AC={enc1[28]}")

    # ── Test 2: PROPIA con 1 jurisdiccion IIBB (CABA) ───────────────
    print("\n=== Test 2: PROPIA con 1 IIBB CABA ===")
    t2 = [{
        "id_operacion": "T2", "fecha": "2026-05-01",
        "usuario": "FABRICIO DAURIA", "oficina": "BUENOS AIRES",
        "numero_carpeta": "EXP-001", "tipo_operacion": "Exportacion",
        "cliente": "Exportadora Y", "concepto": "Honorarios despachante",
        "monto_concepto": 120000, "factura_tipo": "A", "codigo_afip": "001",
        "sucursal": "00003", "numero_factura": "00099001",
        "n_comprobante": "0000300099001", "proveedor_validado": "Si",
        "cuit_proveedor": "20345678901", "cuit_cliente": "30570717630",
        "neto_gravado": 33057.85, "no_gravado": 0,
        "iva_21": 6942.15, "iva_105": 0, "iva_27": 0,
        "perc_iva": 0, "perc_ganancias": 0,
        "perc_iibb": 1000.0, "jurisdiccion": "CABA",
        "perc_iibb_2": 0, "jurisdiccion_iibb_2": "",
        "perc_municipal": 0, "jurisdiccion_municipal": "",
        "monto_total_ticket": 41000.0, "monto_a_imputar": 41000.0,
        "ticket_url": "", "estado": "CERRADO",
        "clave_maestra": "", "observaciones": "",
    }]
    g2 = agrupar_por_comprobante(t2)
    f2 = generar_filas_dux(g2, CUITS_PROPIOS, mock_concepto_fn, mock_empleado_fn)
    enc2 = [r for r in f2 if r[0] == "ENC"][0]
    check("ENC is PROPIA", enc2[1] == "PROPIA")
    check("IIBB slot 1 = (1000, CABA)", enc2[23] == 1000.0 and enc2[24] == "CABA", f"X={enc2[23]} Y={enc2[24]}")
    check("IIBB slot 2 empty", enc2[25] == "" and enc2[26] == "", f"Z={enc2[25]} AA={enc2[26]}")
    check("IIBB slot 3 empty", enc2[27] == "" and enc2[28] == "", f"AB={enc2[27]} AC={enc2[28]}")

    # ���─ Test 3: PROPIA con 2 IIBB (CABA + CORDOBA) ──────���───────────
    print("\n=== Test 3: PROPIA con 2 IIBB (CABA + CORDOBA) ===")
    t3 = [{
        "id_operacion": "T3", "fecha": "2026-05-01",
        "usuario": "FABRICIO DAURIA", "oficina": "BUENOS AIRES",
        "numero_carpeta": "EXP-002", "tipo_operacion": "Exportacion",
        "cliente": "Exportadora Y", "concepto": "Honorarios despachante",
        "monto_concepto": 120000, "factura_tipo": "A", "codigo_afip": "001",
        "sucursal": "00003", "numero_factura": "00099002",
        "n_comprobante": "0000300099002", "proveedor_validado": "Si",
        "cuit_proveedor": "20345678901", "cuit_cliente": "30570717630",
        "neto_gravado": 33057.85, "no_gravado": 0,
        "iva_21": 6942.15, "iva_105": 0, "iva_27": 0,
        "perc_iva": 0, "perc_ganancias": 0,
        "perc_iibb": 4938.0, "jurisdiccion": "CABA",
        "perc_iibb_2": 4938.0, "jurisdiccion_iibb_2": "CORDOBA",
        "perc_municipal": 0, "jurisdiccion_municipal": "",
        "monto_total_ticket": 49876.0, "monto_a_imputar": 49876.0,
        "ticket_url": "", "estado": "CERRADO",
        "clave_maestra": "", "observaciones": "",
    }]
    g3 = agrupar_por_comprobante(t3)
    f3 = generar_filas_dux(g3, CUITS_PROPIOS, mock_concepto_fn, mock_empleado_fn)
    enc3 = [r for r in f3 if r[0] == "ENC"][0]
    check("ENC is PROPIA", enc3[1] == "PROPIA")
    # Sorted: CABA < CORDOBA
    check("IIBB slot 1 = (4938, CABA)", enc3[23] == 4938.0 and enc3[24] == "CABA", f"X={enc3[23]} Y={enc3[24]}")
    check("IIBB slot 2 = (4938, CORDOBA)", enc3[25] == 4938.0 and enc3[26] == "CORDOBA", f"Z={enc3[25]} AA={enc3[26]}")
    check("IIBB slot 3 empty", enc3[27] == "" and enc3[28] == "", f"AB={enc3[27]} AC={enc3[28]}")

    # Validation should warn (not error) about CORDOBA
    errs3, warns3 = validar_rendiciones_para_export(t3, mock_concepto_fn, mock_empleado_fn, CUITS_PROPIOS)
    check("No blocking errors", len(errs3) == 0, f"errors={len(errs3)}")
    check("No warnings (CORDOBA is now valid jurisdiction)",
          len(warns3) == 0,
          f"Warnings: {[w['tipo'] for w in warns3]}")

    # ── Test 4: PROPIA Easy/Cencosud (2 IIBB + Municipal) ───────────
    print("\n=== Test 4: PROPIA Easy (2 IIBB + Municipal CORDOBA) ===")
    t4 = [{
        "id_operacion": "T4_EASY", "fecha": "2026-05-01",
        "usuario": "FABRICIO DAURIA", "oficina": "CORDOBA",
        "numero_carpeta": "IMP-EASY-001", "tipo_operacion": "Importacion",
        "cliente": "Cencosud / Easy", "concepto": "Gastos generales OF. BS.AS.",
        "monto_concepto": 250000, "factura_tipo": "A", "codigo_afip": "001",
        "sucursal": "00091", "numero_factura": "00312456",
        "n_comprobante": "0009100312456", "proveedor_validado": "Si",
        "cuit_proveedor": "30600928235", "cuit_cliente": "30570717630",
        "neto_gravado": 179308.97, "no_gravado": 0,
        "iva_21": 37654.88, "iva_105": 0, "iva_27": 0,
        "perc_iva": 0, "perc_ganancias": 0,
        "perc_iibb": 4938.50, "jurisdiccion": "CORDOBA",
        "perc_iibb_2": 4938.50, "jurisdiccion_iibb_2": "CABA",
        "perc_municipal": 2963.10, "jurisdiccion_municipal": "CORDOBA",
        "monto_total_ticket": 229803.95, "monto_a_imputar": 229803.95,
        "ticket_url": "", "estado": "CERRADO",
        "clave_maestra": "", "observaciones": "",
    }]
    g4 = agrupar_por_comprobante(t4)
    f4 = generar_filas_dux(g4, CUITS_PROPIOS, mock_concepto_fn, mock_empleado_fn)
    enc4 = [r for r in f4 if r[0] == "ENC"][0]
    check("ENC is PROPIA", enc4[1] == "PROPIA")

    # Municipal CORDOBA ($2963.10) goes to No Gravado (confirmed by Fabian 04/05)
    # IIBB: CABA $4938.50 + CORDOBA $4938.50 (only actual IIBB)
    # Sorted: CABA < CORDOBA
    check("IIBB slot 1 = CABA $4938.50", enc4[24] == "CABA" and abs(enc4[23] - 4938.50) < 0.01, f"X={enc4[23]} Y={enc4[24]}")
    check("IIBB slot 2 = CORDOBA $4938.50 (solo IIBB, sin Municipal)", enc4[26] == "CORDOBA" and abs(enc4[25] - 4938.50) < 0.01, f"Z={enc4[25]} AA={enc4[26]}")
    check("IIBB slot 3 empty", enc4[27] == "" and enc4[28] == "", f"AB={enc4[27]} AC={enc4[28]}")
    # No Gravado should include the municipal perception
    check("No Gravado includes Municipal $2963.10",
          abs(enc4[18] - 2963.10) < 0.01,
          f"P(noGravado)={enc4[18]}")

    # ── Test 5: Sumatoria Easy ───────────────────────────────────────
    print("\n=== Test 5: Sumatoria Easy = total ticket ===")
    # netoGravado + noGravado + iva21 + perc_iva + perc_ganancias + iibb_total
    # = 179308.97 + 0 + 37654.88 + 0 + 0 + (4938.50 + 4938.50 + 2963.10)
    # = 179308.97 + 37654.88 + 12840.10 = 229803.95
    total_ticket = 229803.95
    neto = enc4[15]       # O: netoGravado
    no_grav = enc4[18]    # P: noGravado
    iva21 = enc4[20]      # R: iva21
    iva105 = enc4[21]     # S: iva10
    iva27 = enc4[22]      # T: iva27
    perc_iva = enc4[29]   # AA: percepcion iva
    perc_gan = enc4[30]   # AB: percepcion ganancias
    iibb_total = sum(enc4[23 + i*2] for i in range(3)
                     if isinstance(enc4[20 + i*2], (int, float)))

    suma = neto + no_grav + iva21 + iva105 + iva27 + perc_iva + perc_gan + iibb_total
    diff = abs(suma - total_ticket)
    check(f"Suma desglose = total ticket",
          diff < 1.0,
          f"suma={suma:.2f} ticket={total_ticket} diff={diff:.2f}")

    # ── Test 6: "0" jurisdictions don't trigger warnings ───────────
    print("\n=== Test 6: '0' jurisdictions = no warning ===")
    t6 = [{
        "id_operacion": "T6", "fecha": "2026-05-01",
        "usuario": "DAVID REQUELME", "oficina": "BUENOS AIRES",
        "numero_carpeta": "IMP-006", "tipo_operacion": "Importacion",
        "cliente": "Cliente X", "concepto": "Flete terrestre",
        "monto_concepto": 50000, "factura_tipo": "B", "codigo_afip": "006",
        "sucursal": "00010", "numero_factura": "00045678",
        "n_comprobante": "0001000045678", "proveedor_validado": "Si",
        "cuit_proveedor": "30712345678", "cuit_cliente": "",
        "neto_gravado": 0, "no_gravado": 50000.0,
        "iva_21": 0, "iva_105": 0, "iva_27": 0,
        "perc_iva": 0, "perc_ganancias": 0,
        "perc_iibb": 0, "jurisdiccion": "",
        "perc_iibb_2": 0, "jurisdiccion_iibb_2": "0",
        "perc_municipal": 0, "jurisdiccion_municipal": "0",
        "monto_total_ticket": 50000.0, "monto_a_imputar": 50000.0,
        "ticket_url": "", "estado": "CERRADO",
        "clave_maestra": "", "observaciones": "",
    }]
    errs6, warns6 = validar_rendiciones_para_export(t6, mock_concepto_fn, mock_empleado_fn, CUITS_PROPIOS)
    check("No errors", len(errs6) == 0, f"errors={len(errs6)}")
    check("No warnings ('0' filtered)", len(warns6) == 0, f"warnings={len(warns6)}")

    # ── Test 7: real invalid jurisdiction = warning not error ────────
    print("\n=== Test 7: CORDOBA jurisdiction = warning only ===")
    t7 = [{
        **t6[0],
        "id_operacion": "T7",
        "perc_iibb": 1000, "jurisdiccion": "CORDOBA",
        "jurisdiccion_iibb_2": "", "jurisdiccion_municipal": "",
    }]
    errs7, warns7 = validar_rendiciones_para_export(t7, mock_concepto_fn, mock_empleado_fn, CUITS_PROPIOS)
    check("No blocking errors", len(errs7) == 0, f"errors={len(errs7)}")
    check("No warnings (CORDOBA is now valid jurisdiction)",
          len(warns7) == 0,
          len(warns7) == 0)
    print("\n=== Test 8: unmapped concept = blocking error ===")
    t8 = [{
        **t6[0],
        "id_operacion": "T8",
        "concepto": "CONCEPTO_NO_MAPEADO",
        "jurisdiccion_iibb_2": "", "jurisdiccion_municipal": "",
    }]
    errs8, warns8 = validar_rendiciones_para_export(t8, mock_concepto_fn, mock_empleado_fn, CUITS_PROPIOS)
    check("1 blocking error", len(errs8) == 1, f"errors={len(errs8)}")
    check("Error is about concepto", "concepto" in errs8[0]["tipo"].lower() if errs8 else False,
          errs8[0]["tipo"] if errs8 else "")

    # ── Test 9: error + warning combo ────────────────────────────────
    print("\n=== Test 9: error + warning combo ===")
    t9 = [{
        **t6[0],
        "id_operacion": "T9",
        "concepto": "CONCEPTO_NO_MAPEADO",
        "perc_iibb": 1000, "jurisdiccion": "CORDOBA",
        "jurisdiccion_iibb_2": "", "jurisdiccion_municipal": "",
    }]
    errs9, warns9 = validar_rendiciones_para_export(t9, mock_concepto_fn, mock_empleado_fn, CUITS_PROPIOS)
    check("1 blocking error", len(errs9) == 1, f"errors={len(errs9)}")
    check("No jurisdiction warning (CORDOBA valid)", len(warns9) == 0, f"warnings={len(warns9)}")

    # ── Summary ──────────────────────────────────────────────────────
    print(f"\n{'='*50}")
    print(f"  {counts['passed']} PASSED, {counts['failed']} FAILED")
    if counts["failed"] == 0:
        print("  ALL TESTS PASSED")
    else:
        print("  SOME TESTS FAILED")
    print(f"{'='*50}")
    sys.exit(1 if counts["failed"] else 0)
